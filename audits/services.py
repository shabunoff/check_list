from io import BytesIO

from openpyxl import Workbook, load_workbook

from .models import Audit, AuditFieldValue, Checklist, ChecklistAssignment, ChecklistField, Executor


class ImportErrorMessage(Exception):
    pass


def _normalize(value):
    if value is None:
        return ''
    return str(value).strip()


def import_executors_from_workbook(file_obj):
    wb = load_workbook(file_obj)
    ws = wb.active

    created_count = 0
    updated_count = 0

    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        full_name = _normalize(row[0] if len(row) > 0 else '')
        city = _normalize(row[1] if len(row) > 1 else '')

        if not full_name or full_name.lower() == 'фио':
            continue
        if not city:
            raise ImportErrorMessage(f'Строка {idx}: не указана группа')

        _, created = Executor.objects.update_or_create(
            full_name=full_name,
            defaults={'city': city, 'is_active': True},
        )
        if created:
            created_count += 1
        else:
            updated_count += 1

    return {'created': created_count, 'updated': updated_count}


def import_checklist_from_workbook(file_obj, title=None):
    raise ImportErrorMessage('Импорт чек-листов отключен. Чек-листы создаются только вручную через интерфейс.')


def _get_ordered_fields_for_export():
    return list(
        ChecklistField.objects.select_related('item', 'item__section', 'item__section__checklist')
        .order_by('item__section__checklist__id', 'item__section__order', 'item__order', 'order', 'id')
    )


def build_export_matrix_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Аудиты (матрица)'

    fields = _get_ordered_fields_for_export()
    header = ['Респондент', 'Группа', 'Статус респондента', 'Кол-во аудитов', 'Завершено аудитов']

    field_headers = []
    for fld in fields:
        field_headers.append(
            f'[{fld.item.section.checklist.title}] {fld.item.section.title} / {fld.item.text[:40]} / {fld.title}'
        )

    ws.append(header + field_headers)

    audits = list(Audit.objects.select_related('executor', 'checklist').all())
    audit_by_executor_checklist = {(a.executor_id, a.checklist_id): a for a in audits}
    audit_stats = {}
    for a in audits:
        if a.executor_id not in audit_stats:
            audit_stats[a.executor_id] = {'total': 0, 'completed': 0}
        audit_stats[a.executor_id]['total'] += 1
        if a.status == Audit.STATUS_COMPLETED:
            audit_stats[a.executor_id]['completed'] += 1

    values = AuditFieldValue.objects.select_related('audit', 'field').all()
    values_by_audit_field = {(v.audit_id, v.field_id): v.value for v in values}

    for executor in Executor.objects.all().order_by('full_name'):
        stats = audit_stats.get(executor.id, {'total': 0, 'completed': 0})
        row = [
            executor.full_name,
            executor.city,
            'Активный' if executor.is_active else 'Неактивный',
            stats['total'],
            stats['completed'],
        ]

        for fld in fields:
            audit = audit_by_executor_checklist.get((executor.id, fld.item.section.checklist_id))
            if not audit:
                row.append('')
                continue
            row.append(values_by_audit_field.get((audit.id, fld.id), ''))

        ws.append(row)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_export_rows_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Аудиты (по полям)'

    ws.append([
        'Респондент',
        'Группа',
        'Чек-лист',
        'Раздел',
        'Пункт',
        'Поле',
        'Тип поля',
        'Значение',
        'Статус аудита',
    ])

    values = AuditFieldValue.objects.select_related(
        'audit',
        'audit__executor',
        'audit__checklist',
        'field',
        'field__item',
        'field__item__section',
    ).order_by(
        'audit__executor__full_name',
        'field__item__section__order',
        'field__item__order',
        'field__order',
        'id',
    )

    for value in values:
        ws.append([
            value.audit.executor.full_name,
            value.audit.executor.city,
            value.audit.checklist.title,
            value.field.item.section.title,
            value.field.item.text,
            value.field.title,
            value.field.get_field_type_display(),
            value.value,
            value.audit.get_status_display(),
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def build_checklist_assignments_export_workbook(checklist: Checklist):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Назначения'

    ws.append([
        'Чек-лист',
        'Респондент',
        'Группа',
        'Менеджер',
        'Статус назначения',
        'Статус аудита',
        'Заполнено полей',
        'Всего полей',
        'Дата назначения',
    ])

    checklist_field_ids = list(
        ChecklistField.objects.filter(item__section__checklist=checklist).values_list('id', flat=True)
    )
    total_fields = len(checklist_field_ids)

    assignments = ChecklistAssignment.objects.filter(checklist=checklist).select_related('executor', 'manager', 'audit').order_by('executor__full_name')

    for assignment in assignments:
        filled = 0
        audit_status = ''
        if assignment.audit_id:
            audit_status = assignment.audit.get_status_display()
            filled = AuditFieldValue.objects.filter(audit=assignment.audit, field_id__in=checklist_field_ids).exclude(value='').count()

        ws.append([
            checklist.title,
            assignment.executor.full_name,
            assignment.executor.city,
            assignment.manager.username,
            assignment.get_status_display(),
            audit_status,
            filled,
            total_fields,
            assignment.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def build_checklist_answers_export_workbook(checklist: Checklist):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ответы чек-листа'

    ws.append([
        'Чек-лист',
        'Респондент',
        'Группа',
        'Менеджер',
        'Статус назначения',
        'Статус аудита',
        'Раздел',
        'Пункт',
        'Поле',
        'Тип поля',
        'Значение',
    ])

    assignments = (
        ChecklistAssignment.objects.filter(checklist=checklist)
        .select_related('executor', 'manager', 'audit')
        .order_by('executor__full_name')
    )

    fields = list(
        ChecklistField.objects.filter(item__section__checklist=checklist)
        .select_related('item', 'item__section')
        .order_by('item__section__order', 'item__order', 'order', 'id')
    )

    values_qs = AuditFieldValue.objects.filter(
        audit__checklist=checklist,
        field__in=fields,
    ).select_related('audit', 'field')
    values_map = {(v.audit_id, v.field_id): v.value for v in values_qs}

    for assignment in assignments:
        audit_status = assignment.audit.get_status_display() if assignment.audit else ''

        for field in fields:
            value = ''
            if assignment.audit_id:
                value = values_map.get((assignment.audit_id, field.id), '')

            ws.append([
                checklist.title,
                assignment.executor.full_name,
                assignment.executor.city,
                assignment.manager.username,
                assignment.get_status_display(),
                audit_status,
                field.item.section.title,
                field.item.text,
                field.title,
                field.get_field_type_display(),
                value,
            ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

